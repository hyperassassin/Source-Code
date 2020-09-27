from tkinter import *
from openpyxl import *
from tkinter import messagebox , ttk

wb = load_workbook('excel.xlsx') #Full Path can be Specified or keep the file in same folder.
sheet = wb.active

def clear():
    name_field.delete(0,END)
    course_field.delete(0,END)
    sem_field.delete(0,END)
    form_field.delete(0,END)
    contact_field.delete(0,END)
    email_field.delete(0,END)
    address_field.delete(0,END)

def excel():
    #Setting Column Dimensions
    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 30 
    sheet.column_dimensions['H'].width = 30
    sheet.column_dimensions['I'].width = 20
    #Heading Values
    sheet.cell(row = 1 , column = 1).value = " Form No "
    sheet.cell(row = 1 , column = 2).value = " Name "
    sheet.cell(row = 1 , column = 3).value = " Course "
    sheet.cell(row = 1 , column = 4).value = " Semester "
    sheet.cell(row = 1 , column = 5).value = " Gender "
    sheet.cell(row = 1 , column = 6).value = " Date of Birth "
    sheet.cell(row = 1 , column = 7).value = " Email Id "
    sheet.cell(row = 1 , column = 8).value = " Address "
    sheet.cell(row = 1 , column = 9).value = " Contact No "

def insert():
    if(form_field.get() == "" or  name_field.get() == "" or course_field.get() == "" or
       sem_field.get() == "" or combo.get() == "" or c1.get() == "" or c2.get() == "" or c3.get() == "" or
       email_field.get() == "" or address_field.get() == "" or contact_field.get() == ""):
        messagebox.showerror("Python","Fill Out All The Fields")
    else:
        curr_row = sheet.max_row
        curr_col = sheet.max_column
        sheet.cell(row = curr_row + 1 , column = 1).value = form_field.get()
        sheet.cell(row = curr_row + 1 , column = 2).value = name_field.get()
        sheet.cell(row = curr_row + 1 , column = 3).value = course_field.get()
        sheet.cell(row = curr_row + 1 , column = 4).value = sem_field.get()
        sheet.cell(row = curr_row + 1 , column = 5).value = combo.get()
        sheet.cell(row = curr_row + 1 , column = 6).value = c2.get() + "-" + c1.get() + "-" + c3.get()
        sheet.cell(row = curr_row + 1 , column = 7).value = email_field.get()
        sheet.cell(row = curr_row + 1 , column = 8).value = address_field.get()
        sheet.cell(row = curr_row + 1 , column = 9).value = contact_field.get() 
        wb.save('excel.xlsx')
        name_field.focus_set()
        clear()

root = Tk()
root.title(" Registration Form ")
root.geometry("450x530")
root.resizable(0,0)
root.configure(background = "Salmon")

#---------------------Labels------------------------#
form = Label(root,text = " Form No :- ",bg="Salmon").grid(row = 0 , column = 0 , ipadx = 10 , ipady = 10)
name = Label(root,text = " Name :- ",bg="Salmon").grid(row = 1 , column = 0, ipadx = 10 , ipady = 10)
course = Label(root,text = " Course :- ",bg="Salmon").grid(row = 2 , column = 0, ipadx = 10 , ipady = 10)
sem = Label(root,text = " Semester :- ",bg="Salmon").grid(row = 3 , column = 0, ipadx = 10 , ipady = 10)
gender = Label(root,text = " Gender :- ",bg="Salmon").grid(row = 4 , column = 0, ipadx = 10 , ipady = 10)
dob = Label(root,text = " Date Of Birth :- ",bg="Salmon").grid(row = 5 , column = 0, ipadx = 10 , ipady = 10)
email = Label(root,text = " Email  Id :- ",bg="Salmon").grid(row = 6 , column = 0, ipadx = 10 , ipady = 10)
address = Label(root,text = " Address :- ",bg="Salmon").grid(row = 7 , column = 0, ipadx = 10 , ipady = 10)
contact = Label(root,text = " Contact No :- ",bg="Salmon").grid(row = 8 , column = 0, ipadx = 10 , ipady = 10)

#--------------------Entry--------------------#
form_field = Entry(root)
form_field.grid(row = 0 , column = 1)

name_field = Entry(root)
name_field.grid(row = 1 , column = 1)

course_field = Entry(root)
course_field.grid(row = 2 , column = 1)

sem_field = Entry(root)
sem_field.grid(row = 3 , column = 1)

email_field = Entry(root)
email_field.grid(row = 6 , column = 1)

address_field = Entry(root)
address_field.grid(row = 7 , column = 1)
                   
contact_field = Entry(root)
contact_field.grid(row = 8 , column = 1)
                   
excel()
#--------------------Button--------------------#
submit = Button(root , text = " Submit " , command = insert,bg="Salmon").grid(row = 9 , column = 1, ipadx = 10 , ipady = 10)

#---Variables---#
var = StringVar()
var1 = StringVar()
var2 = IntVar()
var3 = IntVar()

#---------------------------Combobox------------------------------#
combo =ttk.Combobox(root,textvariable = var,width = 17)
combo['values'] = ("Male","Female","Other")
combo.grid(row = 4 , column = 1, ipadx = 10 , ipady = 10)

c1 = ttk.Combobox(root , textvariable = var1,width=10)
c1['values'] = ('January','February','March','April','May','June','July'
                   ,'August','September','October','November','December')
c1.place(relx = 0.32 , rely = 0.205)
#For Run In PC Use This :- c1.place(relx = 0.2 , rely = 0.2)

c2 = ttk.Combobox(root , textvariable = var2,width=6)
c2['values'] = ('1','2','3','4','5','6','7','8','9','10','11','12','13','14','15','16','17','18','19','20',
                '21','22','23','24','25','26','27','28','29','30','31')
c2.place(relx = 0.6 , rely = 0.205)
#For Run In PC Use This :- c2.place(relx = 0.34 , rely = 0.2)

c3 = ttk.Combobox(root , textvariable = var3,width=6)
c3['values'] = ('2020','2019','2018','2017','2016','2015','2014','2013','2012','2011','2010',
                '2009','2008','2007','2006','2005','2004','2003','2002','2001','2000','1999','1998',
                '1997','1996','1995','1994','1993','1992','1991','1990','1989','1988','1987','1986',
                '1985','1984','1983','1982','1981','1980')
c3.place(relx= 0.8 , rely = 0.205)
#For Run In PC Use This :- c3.place(relx= 0.48 , rely = 0.2)

root.mainloop()
